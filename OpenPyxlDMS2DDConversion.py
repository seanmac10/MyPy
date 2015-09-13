import os, sys, re, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# Subroutine to convert DMS values like 32° 29' 35.51" N to decimal degrees
#
def dms2dec(dms_str):
    dms_str = re.sub(r'\s', '', dms_str)    
    if re.match('[swSW]', dms_str):
        sign = -1
    else:
        sign = 1
    
    (degree, minute, second, frac_seconds, junk) = re.split('\D+', dms_str, maxsplit=4)
    
    return sign * (int(degree)  + float(minute) / 60 + float(second) / 3600 + float(frac_seconds) / 36000)

# MAIN ----------------------------------------
    
# Load workbook with DMS values
wb = load_workbook(r'C:\temp\DMS2DD.xlsx')

# Select the worksheet
ws = wb.get_sheet_by_name("points")

# Run through selected cells
for row in ws.iter_rows('A2:B12'):
    for cell in row:
        cell.value = dms2dec(cell.value)

# Save converted workbook        
wb.save(r'C:\temp\dmsconverted.xlsx')


